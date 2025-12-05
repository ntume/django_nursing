import csv
from io import StringIO
from django.shortcuts import render, redirect
from django.urls import reverse_lazy
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.generic import ListView
from django.http import JsonResponse
from openpyxl import load_workbook
from io import BytesIO

from .models import *
from .forms import *

# Create your views here.


class NationalityListView(LoginRequiredMixin,ListView):
    template_name = 'configurable/nationalities.html'
    context_object_name = 'items'
    model = Nationality

    def get_queryset(self):
        return Nationality.objects.all()

    def get_context_data(self, **kwargs):
        context = super(NationalityListView, self).get_context_data(**kwargs) 
        context['config_menu'] = '--active'
        context['config_menu_open'] = 'side-menu__sub-open'
        return context
    

@login_required()
def add_nationality(request):

    form = NationalityForm(request.POST)
    if form.is_valid():       
        form.save()
        messages.success(request,'Successfully added Nationality')
    else:
        messages.warning(request,form.errors)

    return redirect('administration:nationalities')

@login_required()
def edit_nationality(request,pk):

    item = Nationality.objects.get(id=pk)
    form = NationalityForm(request.POST,instance=item)

    if form.is_valid():
        form.save()
        messages.success(request,'Successfully edited Nationality')
    else:
        messages.warning(request,form.errors)

    return redirect('administration:nationalities')

@login_required()
def delete_nationality(request,pk):

    try:
        item = Nationality.objects.get(id=pk)
        item.delete()
        messages.success(request, 'Successfully deleted nationality')
    except Exception as e:
        messages.warning(request, str(e))

    return redirect('administration:nationalities')

class StructureStatusIDListView(LoginRequiredMixin,ListView):
    template_name = 'configurable/structure_status_id.html'
    context_object_name = 'statuses'
    model = StructureStatusID

    def get_queryset(self):
        return StructureStatusID.objects.all()

    def get_context_data(self, **kwargs):
        context = super(StructureStatusIDListView, self).get_context_data(**kwargs) 
        context['config_menu'] = '--active'
        context['config_menu_open'] = 'side-menu__sub-open'
        return context
    

@login_required()
def add_structure_status_id(request):

    form = StructureStatusIDForm(request.POST)
    if form.is_valid():       
        form.save()
        messages.success(request,'Successfully added status')
    else:
        messages.warning(request,form.errors)

    return redirect('administration:stucture_status_ids')

@login_required()
def edit_structure_status_id(request,pk):

    type_of_id = StructureStatusID.objects.get(id=pk)
    form = StructureStatusIDForm(request.POST,instance=type_of_id)

    if form.is_valid():
        form.save()
        messages.success(request,'Successfully edited status')
    else:
        messages.warning(request,form.errors)

    return redirect('administration:stucture_status_ids')

@login_required()
def delete_structure_status_id(request,pk):

    try:
        type_of_id_instance = StructureStatusID.objects.get(id=pk)
        type_of_id_instance.delete()
        messages.success(request, 'Successfully deleted status')
    except Exception as e:
        messages.warning(request, str(e))

    return redirect('administration:stucture_status_ids')



class TypeOfIDListView(LoginRequiredMixin,ListView):
    template_name = 'configurable/type_of_id.html'
    context_object_name = 'types'
    model = TypeOfID

    def get_queryset(self):
        return TypeOfID.objects.all()

    def get_context_data(self, **kwargs):
        context = super(TypeOfIDListView, self).get_context_data(**kwargs) 
        context['config_menu'] = '--active'
        context['config_menu_open'] = 'side-menu__sub-open'
        context['type_of_id_menu_open'] = '--active'
        return context
    

@login_required()
def add_type_of_id(request):

    form = TypeOfIDForm(request.POST)
    if form.is_valid():       
        form.save()
        messages.success(request,'Successfully added type of ID')
    else:
        messages.warning(request,form.errors)

    return redirect('administration:type_of_ids')

@login_required()
def edit_type_of_id(request,pk):

    type_of_id = TypeOfID.objects.get(id=pk)
    form = TypeOfIDForm(request.POST,instance=type_of_id)

    if form.is_valid():
        form.save()
        messages.success(request,'Successfully edited type of ID')
    else:
        messages.warning(request,form.errors)

    return redirect('administration:type_of_ids')

@login_required()
def delete_type_of_id(request,pk):

    try:
        type_of_id_instance = TypeOfID.objects.get(id=pk)
        type_of_id_instance.delete()
        messages.success(request, 'Successfully deleted type of ID')
    except Exception as e:
        messages.warning(request, str(e))

    return redirect('administration:type_of_ids')



class CountryListView(LoginRequiredMixin,ListView):
    template_name = 'configurable/countries.html'
    context_object_name = 'countries'
    model = Country

    def get_queryset(self):
        return Country.objects.all()

    def get_context_data(self, **kwargs):
        context = super(CountryListView, self).get_context_data(**kwargs) 
        context['config_menu'] = '--active'
        context['config_menu_open'] = 'side-menu__sub-open'
        context['country_menu_open'] = '--active'
        return context
    

@login_required()
def add_country(request):

    form = CountryForm(request.POST)
    if form.is_valid():       
        form.save()
        messages.success(request,'Successfully added country')
    else:
        messages.warning(request,form.errors)

    return redirect('administration:countries')

@login_required()
def edit_country(request,pk):

    country = Country.objects.get(id=pk)
    form = CountryForm(request.POST,instance=country)

    if form.is_valid():
        form.save()
        messages.success(request,'Successfully edited country')
    else:
        messages.warning(request,form.errors)

    return redirect('administration:countries')

@login_required()
def delete_country(request,pk):

    try:
        country_instance = Country.objects.get(id=pk)
        country_instance.delete()
        messages.success(request, 'Successfully deleted country')
    except Exception as e:
        messages.warning(request, str(e))

    return redirect('administration:countries')

    
class ProvinceListView(LoginRequiredMixin,ListView):
    template_name = 'configurable/provinces.html'
    context_object_name = 'provinces'
    model = Province
    paginate_by = 50

    def get_queryset(self):
        return Province.objects.filter(country_id=self.kwargs['pk'])

    def get_context_data(self, **kwargs):
        context = super(ProvinceListView, self).get_context_data(**kwargs) 
        context['country'] = Country.objects.get(id = self.kwargs['pk'])
        context['config_menu'] = '--active'
        context['config_menu_open'] = 'side-menu__sub-open'
        context['province_menu_open'] = '--active'
        return context
    

@login_required()
def add_province(request,country_pk):

    form = ProvinceForm(request.POST)
    if form.is_valid():
        province = form.save(commit=False)
        province.country_id = country_pk
        province.save()
        messages.success(request,'Successfully added province')
    else:
        messages.warning(request,form.errors)

    return redirect('administration:provinces',pk=country_pk)

@login_required()
def edit_province(request,pk):

    province_instance = Province.objects.get(id=pk)
    form = ProvinceForm(request.POST,instance=province_instance)

    if form.is_valid():
        form.save()
        messages.success(request,'Successfully edited province')
    else:
        messages.warning(request,form.errors)

    return redirect('administration:provinces',pk=province_instance.country_id)

@login_required()
def delete_province(request,pk):

    try:
        province_instance = Province.objects.get(id=pk)
        country_pk = province_instance.country_id
        province_instance.delete()
        messages.success(request, 'Successfully deleted province')
    except Exception as e:
        messages.warning(request, str(e))

    return redirect('administration:provinces',pk=country_pk)



class MunicipalityListView(LoginRequiredMixin,ListView):
    template_name = 'configurable/municipalities.html'
    context_object_name = 'municipalities'
    model = Municipality

    def get_queryset(self):
        return Municipality.objects.filter(province_id=self.kwargs['pk'])

    def get_context_data(self, **kwargs):
        context = super(MunicipalityListView, self).get_context_data(**kwargs) 
        context['province'] = Province.objects.get(id = self.kwargs['pk'])
        context['config_menu'] = '--active'
        context['config_menu_open'] = 'side-menu__sub-open'
        context['municipality_menu_open'] = '--active'
        return context
    

@login_required()
def add_municipality(request,province_pk):

    form = MunicipalityForm(request.POST)
    if form.is_valid():
        municipality = form.save(commit=False)
        municipality.province_id = province_pk
        municipality.save()
        messages.success(request,'Successfully added municipality')
    else:
        messages.warning(request,form.errors)

    return redirect('administration:municipalities',pk=province_pk)

@login_required()
def edit_municipality(request,pk):

    municipality_instance = Municipality.objects.get(id=pk)
    form = MunicipalityForm(request.POST,instance=municipality_instance)

    if form.is_valid():
        form.save()
        messages.success(request,'Successfully edited municipality')
    else:
        messages.warning(request,form.errors)

    return redirect('administration:municipalities',pk=municipality_instance.province_id)

@login_required()
def delete_municipality(request,pk):

    try:
        municipality_instance = Municipality.objects.get(id=pk)
        province_pk = municipality_instance.province_id
        municipality_instance.delete()
        messages.success(request, 'Successfully deleted municipality')
    except Exception as e:
        messages.warning(request, str(e))

    return redirect('administration:municipalities',pk=province_pk)



class DistrictListView(LoginRequiredMixin,ListView):
    template_name = 'configurable/districts.html'
    context_object_name = 'districts'
    model = Municipality

    def get_queryset(self):
        return District.objects.filter(municipality_id=self.kwargs['pk'])

    def get_context_data(self, **kwargs):
        context = super(DistrictListView, self).get_context_data(**kwargs) 
        context['municipality'] = Municipality.objects.get(id = self.kwargs['pk'])
        context['config_menu'] = '--active'
        context['config_menu_open'] = 'side-menu__sub-open'
        context['district_menu_open'] = '--active'
        return context
    

@login_required()
def add_district(request,municipality_pk):

    form = DistrictForm(request.POST)
    if form.is_valid():
        district = form.save(commit=False)
        district.municipality_id = municipality_pk
        district.save()
        messages.success(request,'Successfully added district')
    else:
        messages.warning(request,form.errors)

    return redirect('administration:districts',pk=municipality_pk)

@login_required()
def edit_district(request,pk):

    district_instance = District.objects.get(id=pk)
    form = DistrictForm(request.POST,instance=district_instance)

    if form.is_valid():
        form.save()
        messages.success(request,'Successfully edited district')
    else:
        messages.warning(request,form.errors)

    return redirect('administration:districts',pk=district_instance.municipality_id)

@login_required()
def delete_district(request,pk):

    try:
        district_instance = District.objects.get(id=pk)
        municipality_pk = District.municipality_id
        district_instance.delete()
        messages.success(request, 'Successfully deleted district')
    except Exception as e:
        messages.warning(request, str(e))

    return redirect('administration:districts',pk=municipality_pk)


class CityList(LoginRequiredMixin,ListView):
    template_name = 'configurable/cities.html'
    context_object_name = 'cities'
    model = City

    def get_queryset(self):
        return City.objects.filter(district_id=self.kwargs['pk'])

    def get_context_data(self, **kwargs):
        context = super(CityList, self).get_context_data(**kwargs) 
        context['district'] = District.objects.get(id = self.kwargs['pk'])
        context['config_menu'] = '--active'
        context['config_menu_open'] = 'side-menu__sub-open'
        context['city_menu_open'] = '--active'
        return context


@login_required()
def add_city(request,district_pk):

    form = CityForm(request.POST)
    if form.is_valid():
        city = form.save(commit=False)
        city.district_id = district_pk
        city.save()
        messages.success(request,'Successfully added City')
    else:
        messages.warning(request,form.errors)

    return redirect('administration:cities',pk=district_pk)


@login_required()
def edit_city(request,pk):

    city_instance = City.objects.get(id = pk)
    form = CityForm(request.POST,instance=city_instance)
    if form.is_valid():
        form.save()
        messages.success(request,'Successfully added City')
    else:
        messages.warning(request,form.errors)

    return redirect('administration:cities',pk=city_instance.district_id)


@login_required()
def delete_city(request,pk):

    try:
        city_instance = City.objects.get(id=pk)
        district_pk = city_instance.district_id
        city_instance.delete()
        messages.success(request, 'Successfully deleted city')
    except Exception as e:
        messages.warning(request, str(e))

    return redirect('administration:cities',pk=district_pk)



class SuburbList(LoginRequiredMixin,ListView):
    template_name = 'configurable/suburbs.html'
    context_object_name = 'suburbs'
    model = Suburb

    def get_queryset(self):
        return Suburb.objects.filter(city_id=self.kwargs['pk'])

    def get_context_data(self, **kwargs):
        context = super(SuburbList, self).get_context_data(**kwargs) 
        context['city'] = City.objects.get(id = self.kwargs['pk'])
        context['config_menu'] = '--active'
        context['config_menu_open'] = 'side-menu__sub-open'
        context['suburb_menu_open'] = '--active'
        return context


@login_required()
def add_suburb(request,city_pk):

    form = SuburbForm(request.POST)
    if form.is_valid():
        suburb = form.save(commit=False)
        suburb.city_id = city_pk
        suburb.save()
        messages.success(request,'Successfully added Suburb')
    else:
        messages.warning(request,form.errors)

    return redirect('administration:suburbs',pk=city_pk)


@login_required()
def edit_suburb(request,pk):

    suburb_instance = Suburb.objects.get(id = pk)
    form = SuburbForm(request.POST,instance=suburb_instance)
    if form.is_valid():
        form.save()
        messages.success(request,'Successfully added suburb')
    else:
        messages.warning(request,form.errors)

    return redirect('administration:suburbs',pk=suburb_instance.city_id)


@login_required()
def delete_suburb(request,pk):

    try:
        suburb_instance = City.objects.get(id=pk)
        city_pk = suburb_instance.city_id
        suburb_instance.delete()
        messages.success(request, 'Successfully deleted suburb')
    except Exception as e:
        messages.warning(request, str(e))

    return redirect('administration:suburbs',pk=city_pk)



def populate_municipality(request):
    '''
    function to populate province
    '''

    districts = District.objects.all()
    for d in districts:
        d.municipality_id = d.id
        d.save()


class LanguageList(LoginRequiredMixin,ListView):
    template_name = 'configurable/languages.html'
    context_object_name = 'languages'
    model = Language

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['config_menu'] = '--active'
        context['config_menu_open'] = 'side-menu__sub-open'
        context['languag_menu_open'] = '--active'
        return context


@login_required()
def add_language(request):

    form = LanguageForm(request.POST)
    if form.is_valid():
        language = form.save()
        messages.success(request,'Successfully added language')
    else:
        messages.warning(request,form.errors)

    return redirect('administration:languages')

@login_required()
def edit_language(request,pk):

    language_instance = Language.objects.get(id=pk)
    form = LanguageForm(request.POST,instance=language_instance)

    if form.is_valid():
        language = form.save(commit=True)
        messages.success(request,'Successfully edited language')
    else:
        messages.warning(request,form.errors)

    return redirect('administration:languages')

@login_required()
def delete_language(request,pk):

    try:
        language_instance = Language.objects.get(id=pk)
        language_instance.delete()
        messages.success(request, 'Successfully deleted language')
    except Exception as e:
        messages.warning(request, str(e))

    return redirect('administration:languages')


class GenderList(LoginRequiredMixin,ListView):
    template_name = 'configurable/gender.html'
    context_object_name = 'gender'
    model = Gender

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['config_menu'] = '--active'
        context['config_menu_open'] = 'side-menu__sub-open'
        context['gender_menu_open'] = '--active'
        return context


@login_required()
def add_gender(request):

    form = GenderForm(request.POST)
    if form.is_valid():
        gender = form.save()
        messages.success(request,'Successfully added gender')
    else:
        messages.warning(request,form.errors)

    return redirect('administration:gender')

@login_required()
def edit_gender(request,pk):

    gender_instance = Gender.objects.get(id=pk)
    form = GenderForm(request.POST,instance=gender_instance)

    if form.is_valid():
        gender = form.save(commit=True)
        messages.success(request,'Successfully edited gender')
    else:
        messages.warning(request,form.errors)

    return redirect('administration:gender')

@login_required()
def delete_gender(request,pk):

    try:
        gender_instance = Gender.objects.get(id=pk)
        gender_instance.delete()
        messages.success(request, 'Successfully deleted gender')
    except Exception as e:
        messages.warning(request, str(e))

    return redirect('administration:gender')


class EconomicStatusList(LoginRequiredMixin,ListView):
    template_name = 'configurable/economic_status.html'
    context_object_name = 'status'
    model = EconomicStatus

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['config_menu'] = '--active'
        context['config_menu_open'] = 'side-menu__sub-open'
        context['economic_status_menu_open'] = '--active'
        return context


@login_required()
def add_economic_status(request):

    form = EconomicStatusForm(request.POST)
    if form.is_valid():
        form.save()
        messages.success(request,'Successfully added Economic Status')
    else:
        messages.warning(request,form.errors)

    return redirect('administration:economic_status')

@login_required()
def edit_economic_status(request,pk):

    ec_instance = EconomicStatus.objects.get(id=pk)
    form = EconomicStatusForm(request.POST,instance=ec_instance)

    if form.is_valid():
        form.save()
        messages.success(request,'Successfully edited Economic Status')
    else:
        messages.warning(request,form.errors)

    return redirect('administration:economic_status')

@login_required()
def delete_economic_status(request,pk):

    try:
        ec_instance = EconomicStatus.objects.get(id=pk)
        ec_instance.delete()
        messages.success(request, 'Successfully deleted Economic Status')
    except Exception as e:
        messages.warning(request, str(e))

    return redirect('administration:economic_status')


class EmploymentEconomicStatusList(LoginRequiredMixin,ListView):
    template_name = 'configurable/employment_economic_status.html'
    context_object_name = 'status'
    model = EmploymentEconomicStatus

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['config_menu'] = '--active'
        context['config_menu_open'] = 'side-menu__sub-open'
        context['economic_status_menu_open'] = '--active'
        return context


@login_required()
def add_employment_economic_status(request):

    form = EmploymentEconomicStatusForm(request.POST)
    if form.is_valid():
        form.save()
        messages.success(request,'Successfully added Economic Status')
    else:
        messages.warning(request,form.errors)

    return redirect('administration:employment_economic_status')

@login_required()
def edit_employment_economic_status(request,pk):

    ec_instance = EmploymentEconomicStatus.objects.get(id=pk)
    form = EmploymentEconomicStatusForm(request.POST,instance=ec_instance)

    if form.is_valid():
        form.save()
        messages.success(request,'Successfully edited Economic Status')
    else:
        messages.warning(request,form.errors)

    return redirect('administration:employment_economic_status')

@login_required()
def delete_employment_economic_status(request,pk):

    try:
        ec_instance = EmploymentEconomicStatus.objects.get(id=pk)
        ec_instance.delete()
        messages.success(request, 'Successfully deleted Economic Status')
    except Exception as e:
        messages.warning(request, str(e))

    return redirect('administration:employment_economic_status')


class RaceList(LoginRequiredMixin,ListView):
    template_name = 'configurable/race.html'
    context_object_name = 'races'
    model = Race

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['config_menu'] = '--active'
        context['config_menu_open'] = 'side-menu__sub-open'
        context['race_menu_open'] = '--active'
        return context


@login_required()
def add_race(request):

    form = RaceForm(request.POST)
    if form.is_valid():
        form.save()
        messages.success(request,'Successfully added race')
    else:
        messages.warning(request,form.errors)

    return redirect('administration:races')

@login_required()
def edit_race(request,pk):

    race_instance = Race.objects.get(id=pk)
    form = RaceForm(request.POST,instance=race_instance)

    if form.is_valid():
        form.save()
        messages.success(request,'Successfully edited race')
    else:
        messages.warning(request,form.errors)

    return redirect('administration:races')

@login_required()
def delete_race(request,pk):

    try:
        race_instance = Race.objects.get(id=pk)
        race_instance.delete()
        messages.success(request, 'Successfully deleted race')
    except Exception as e:
        messages.warning(request, str(e))

    return redirect('administration:races')


class ResidentialStatusList(LoginRequiredMixin,ListView):
    template_name = 'configurable/residential_status.html'
    context_object_name = 'status'
    model = ResidentialStatus

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['config_menu'] = '--active'
        context['config_menu_open'] = 'side-menu__sub-open'
        context['residential_menu_open'] = '--active'
        return context


@login_required()
def add_residential_status(request):

    form = ResidentialStatusForm(request.POST)
    if form.is_valid():
        form.save()
        messages.success(request,'Successfully added residential Status')
    else:
        messages.warning(request,form.errors)

    return redirect('administration:residential_status')

@login_required()
def edit_residential_status(request,pk):

    r_instance = ResidentialStatus.objects.get(id=pk)
    form = ResidentialStatusForm(request.POST,instance=r_instance)

    if form.is_valid():
        form.save()
        messages.success(request,'Successfully edited residential Status')
    else:
        messages.warning(request,form.errors)

    return redirect('administration:residential_status')

@login_required()
def delete_residential_status(request,pk):

    try:
        r_instance = ResidentialStatus.objects.get(id=pk)
        r_instance.delete()
        messages.success(request, 'Successfully deleted residential Status')
    except Exception as e:
        messages.warning(request, str(e))

    return redirect('administration:residential_status')


class DisabilityList(LoginRequiredMixin,ListView):
    template_name = 'configurable/disabilities.html'
    context_object_name = 'disabilities'
    model = Disability

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['config_menu'] = '--active'
        context['config_menu_open'] = 'side-menu__sub-open'
        context['disability_menu_open'] = '--active'
        return context


@login_required()
def add_disability(request):

    form = DisabilityForm(request.POST)
    if form.is_valid():
        form.save()
        messages.success(request,'Successfully added disability')
    else:
        messages.warning(request,form.errors)

    return redirect('administration:disabilities')

@login_required()
def edit_disability(request,pk):

    d_instance = Disability.objects.get(id=pk)
    form = DisabilityForm(request.POST,instance=d_instance)

    if form.is_valid():
        form.save()
        messages.success(request,'Successfully edited disability')
    else:
        messages.warning(request,form.errors)

    return redirect('administration:disabilities')

@login_required()
def delete_disability(request,pk):

    try:
        d_instance = Disability.objects.get(id=pk)
        d_instance.delete()
        messages.success(request, 'Successfully deleted disability')
    except Exception as e:
        messages.warning(request, str(e))

    return redirect('administration:disabilities')


class QuestionTypeList(LoginRequiredMixin,ListView):
    template_name = 'configurable/question_types.html'
    context_object_name = 'question_types'
    model = QuestionType

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['config_menu'] = '--active'
        context['config_menu_open'] = 'side-menu__sub-open'
        context['question_type_menu_open'] = '--active'
        return context

@login_required()
def add_question_types(request):

    if request.method == "POST":
        form = QuestionTypeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request,"Successfully added question type")
        else:
            messages.warning(request,form.errors)

    return redirect('administration:question_types')

@login_required()
def delete_question_type(request,pk):
    try:
        type_instance = QuestionType.objects.get(id=pk)
        type_instance.delete()
        messages.success(request,'Successfully deleted Question Type')
    except:
        messages.warning(request,"An error has occurred")

    return redirect('administration:question_types')

@login_required()
def edit_question_type(request,pk):
    type_instance = QuestionType.objects.get(id=pk)
    form = QuestionTypeForm(request.POST,instance=type_instance)
    if form.is_valid():
        form.save()
        messages.success(request,"Successfully edited question type")
    else:
        messages.warning(request,form.errors)

    return redirect('administration:question_types')

@login_required()
def question_type_add_option(request,pk):
    type_instance = QuestionType.objects.get(id=pk)
    form = QuestionTypeOptionsForm(request.POST)
    if form.is_valid():
        options = form.save(commit=False)
        options.type = type_instance
        options.save()
        messages.success(request,"Successfully added option")
    else:
        messages.warning(request,form.errors)

    return redirect('administration:question_types')

@login_required()
def question_types_delete_option(request,pk):
    try:
        option_instance = QuestionTypeOptions.objects.get(id=pk)
        option_instance.delete()
        messages.success(request,"Successfully removed option")
    except:
        messages.warning(request,"An error has occurred")

    return redirect('administration:question_types')



def ajax_fetch_municipalities(request):
    province_id = request.GET.get('province_id', None)
 
    try:
        province_check = Province.objects.filter(id = province_id)

        if province_check.exists():
            province = province_check.first()
          
            data = {
                'valid':1,
                'province_id':province.id,
            }
            info = []
            for x in province.municipalities.all():
                info.append({'id':x.id,'municipality':x.municipality})

            data['info'] = info
        else:
            data = {
                'valid':0,
                'message':'No Municipaliies'
            }
    except Exception as e:
        data = {
            'valid':2,
            'message':'str(e)'
        }

    return JsonResponse(data)



def ajax_fetch_cities(request):
    municipality_id = request.GET.get('municipality_id', None)
    try:
        municipality_check = Municipality.objects.filter(id = municipality_id)

        if municipality_check.exists():
            municipality = municipality_check.first()
            data = {
                'valid':1,
                'municipality_id':municipality.id,
            }
            info = City.objects.filter(district__municipality = municipality).values('id','city')
            
            data['info'] = list(info)
           
        else:
            data = {
                'valid':0,
                'message':'No Cities'
            }
    except Exception as e:
        data = {
            'valid':2,
            'message':'str(e)'
        }

    return JsonResponse(data)



def ajax_fetch_surburbs(request):
    city_id = request.GET.get('city_id', None)
    try:
        city_check = City.objects.filter(id = city_id)

        if city_check.exists():
            city = city_check.first()
            data = {
                'valid':1,
                'city_id':city.id,
            }
            info = []
            for s in city.suburbs.all():
                info.append({'id':s.id,'text':f'{s.postal_code} {s.suburb}'})

            data['info'] = info
        else:
            data = {
                'valid':0,
                'message':'No Surburbs'
            }
    except Exception as e:
        data = {
            'valid':2,
            'message':'str(e)'
        }

    return JsonResponse(data)



def ajax_verify_area(request):
    area = request.GET.get('area', None)
    area_list = area.split('*')
  
    try:
        check_area = Suburb.objects.filter(suburb = area_list[1].strip(),postal_code = area_list[0].strip())

        if check_area.exists():
            data = {
                'valid':0,
            }
        else:            
            data = {
                'valid':1,
            }
    except Exception as e:
     
        data = {
            'valid':2,
            'message':'Invalid Area, please try again'
        }

    return JsonResponse(data)


def ajax_fetech_areas(request):
    '''
    fetch suburbs for select2
    '''

    term = request.GET.get('term')
    suburbs = Suburb.objects.all().filter(suburb__icontains=term)
    return JsonResponse(list(suburbs.values()), safe=False)


class LeaveList(LoginRequiredMixin,ListView):
    template_name = 'configurable/leave.html'
    context_object_name = 'items'

    def get(self, *args, **kwargs):
        if self.request.user.logged_in_role_id != 2 and self.request.user.logged_in_role_id != 1:
            messages.warning(self.request,"You do not have rights to that portion of the site, you have been logged off!")
            return redirect('accounts:logout')
        return super(LeaveList, self).get(*args, **kwargs)

    def get_queryset(self):
        return  TypeOfLeave.objects.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['config_menu'] = '--active'
        context['config_menu_open'] = 'side-menu__sub-open'
        context['type_of_leave_menu_open'] = '--active'
        return context

@login_required()
def type_of_leave_add(request):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        form = TypeOfLeaveForm(request.POST)
        if form.is_valid():
            form.save()

            messages.success(request,"Successfully added type of leave")
        else:
            messages.warning(request,form.errors)

        return redirect('configurable:type_of_leave')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')

@login_required()
def type_of_leave_edit(request,pk):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        leave_instance = TypeOfLeave.objects.get(id = pk)
        form = TypeOfLeaveForm(request.POST,instance = leave_instance)
        if form.is_valid():
            form.save()
            
            messages.success(request,"Successfully edited type of leave")
        else:
            messages.warning(request,form.errors)

        return redirect('configurable:type_of_leave')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')

@login_required()
def type_of_leave_delete(request,pk):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        try:
            leave_instance = TypeOfLeave.objects.get(id = pk)
            leave_instance.delete()
            messages.success(request,"Successfully removed type of leave")
        except Exception as e:
            messages.warning(request,f"An error has occurred, please try again - Error: {str(e)}")

        return redirect('configurable:type_of_leave')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')
    

class IndemnityList(LoginRequiredMixin,ListView):
    template_name = 'configurable/indemnity.html'
    context_object_name = 'items'

    def get(self, *args, **kwargs):
        if self.request.user.logged_in_role_id != 2 and self.request.user.logged_in_role_id != 1:
            messages.warning(self.request,"You do not have rights to that portion of the site, you have been logged off!")
            return redirect('accounts:logout')
        return super(IndemnityList, self).get(*args, **kwargs)

    def get_queryset(self):
        return  Indemnity.objects.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['config_menu'] = '--active'
        context['config_menu_open'] = 'side-menu__sub-open'
        context['indemnity_menu_open'] = '--active'
        return context

@login_required()
def indemnity_add(request):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        form = IndemnityForm(request.POST)
        if form.is_valid():
            form.save()
            

            messages.success(request,"Successfully added indemnity details")
        else:
            messages.warning(request,form.errors)

        return redirect('configurable:indemnity_list')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')

@login_required()
def indemnity_edit(request,pk):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        staff_instance = Indemnity.objects.get(id = pk)
        form = IndemnityForm(request.POST,instance = staff_instance)
        if form.is_valid():
            form.save()
            
            messages.success(request,"Successfully edited indemnity")
        else:
            messages.warning(request,form.errors)

        return redirect('configurable:indemnity_list')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')

@login_required()
def indemnity_delete(request,pk):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        try:
            i_instance = Indemnity.objects.get(id = pk)
            i_instance.delete()
            messages.success(request,"Successfully deleted Indemnity")
        except Exception as e:
            messages.warning(request,f"An error has occurred, please try again - Error: {str(e)}")

        return redirect('configurable:indemnity_list')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')
    

class SponsorshipList(LoginRequiredMixin,ListView):
    template_name = 'configurable/sponsorship.html'
    context_object_name = 'items'

    def get(self, *args, **kwargs):
        if self.request.user.logged_in_role_id != 2 and self.request.user.logged_in_role_id != 1:
            messages.warning(self.request,"You do not have rights to that portion of the site, you have been logged off!")
            return redirect('accounts:logout')
        return super(SponsorshipList, self).get(*args, **kwargs)

    def get_queryset(self):
        return  Sponsorship.objects.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['config_menu'] = '--active'
        context['config_menu_open'] = 'side-menu__sub-open'
        context['sponsorship_menu_open'] = '--active'
        return context

@login_required()
def sponsorship_add(request):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        form = SponsorshipForm(request.POST)
        if form.is_valid():
            form.save()
            

            messages.success(request,"Successfully added sponsorship details")
        else:
            messages.warning(request,form.errors)

        return redirect('configurable:sponsorship_list')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')

@login_required()
def sponsorship_edit(request,pk):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        staff_instance = Sponsorship.objects.get(id = pk)
        form = SponsorshipForm(request.POST,instance = staff_instance)
        if form.is_valid():
            form.save()
            
            messages.success(request,"Successfully edited sponsorship")
        else:
            messages.warning(request,form.errors)

        return redirect('configurable:sponsorship_list')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')

@login_required()
def sponsorship_delete(request,pk):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        try:
            i_instance = Sponsorship.objects.get(id = pk)
            i_instance.delete()
            messages.success(request,"Successfully deleted Sponsorship")
        except Exception as e:
            messages.warning(request,f"An error has occurred, please try again - Error: {str(e)}")

        return redirect('configurable:sponsorship_list')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')
    
    
class BlockList(LoginRequiredMixin,ListView):
    template_name = 'configurable/blocks.html'
    context_object_name = 'items'

    def get(self, *args, **kwargs):
        if self.request.user.logged_in_role_id != 2 and self.request.user.logged_in_role_id != 1:
            messages.warning(self.request,"You do not have rights to that portion of the site, you have been logged off!")
            return redirect('accounts:logout')
        return super(BlockList, self).get(*args, **kwargs)

    def get_queryset(self):
        return  ProgarmmeBlock.objects.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['config_menu'] = '--active'
        context['config_menu_open'] = 'side-menu__sub-open'
        context['block_menu_open'] = '--active'
        return context

@login_required()
def block_add(request):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        form = ProgarmmeBlockForm(request.POST)
        if form.is_valid():
            form.save()                       

            messages.success(request,"Successfully added Progarmme Block details")
        else:
            messages.warning(request,form.errors)

        return redirect('configurable:block_list')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')

@login_required()
def block_edit(request,pk):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        block_instance = ProgarmmeBlock.objects.get(id = pk)
        form = ProgarmmeBlockForm(request.POST,instance = block_instance)
        if form.is_valid():
            form.save()
            
            messages.success(request,"Successfully edited Progarmme Block")
        else:
            messages.warning(request,form.errors)

        return redirect('configurable:block_list')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')

@login_required()
def block_delete(request,pk):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        try:
            i_instance = ProgarmmeBlock.objects.get(id = pk)
            i_instance.delete()
            messages.success(request,"Successfully deleted Progarmme Block")
        except Exception as e:
            messages.warning(request,f"An error has occurred, please try again - Error: {str(e)}")

        return redirect('configurable:block_list')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')
    


class DemonstrationList(LoginRequiredMixin,ListView):
    template_name = 'configurable/demonstrations.html'
    context_object_name = 'items'

    def get(self, *args, **kwargs):
        if self.request.user.logged_in_role_id != 2 and self.request.user.logged_in_role_id != 1:
            messages.warning(self.request,"You do not have rights to that portion of the site, you have been logged off!")
            return redirect('accounts:logout')
        return super(DemonstrationList, self).get(*args, **kwargs)

    def get_queryset(self):
        return  Demonstration.objects.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['config_menu'] = '--active'
        context['config_menu_open'] = 'side-menu__sub-open'
        context['demonstration_menu_open'] = '--active'
        return context

@login_required()
def demonstration_add(request):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        form = DemonstrationForm(request.POST)
        if form.is_valid():
            form.save()                       

            messages.success(request,"Successfully added Demonstration details")
        else:
            messages.warning(request,form.errors)

        return redirect('configurable:demonstration_list')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')

@login_required()
def demonstration_edit(request,pk):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        block_instance = Demonstration.objects.get(id = pk)
        form = DemonstrationForm(request.POST,instance = block_instance)
        if form.is_valid():
            form.save()
            
            messages.success(request,"Successfully edited Demonstration")
        else:
            messages.warning(request,form.errors)

        return redirect('configurable:demonstration_list')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')

@login_required()
def demonstration_delete(request,pk):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        try:
            i_instance = Demonstration.objects.get(id = pk)
            i_instance.delete()
            messages.success(request,"Successfully deleted Demonstration")
        except Exception as e:
            messages.warning(request,f"An error has occurred, please try again - Error: {str(e)}")

        return redirect('configurable:demonstration_list')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')
    


class DisciplineList(LoginRequiredMixin,ListView):
    template_name = 'configurable/disciplines.html'
    context_object_name = 'items'

    def get(self, *args, **kwargs):
        if self.request.user.logged_in_role_id != 2 and self.request.user.logged_in_role_id != 1:
            messages.warning(self.request,"You do not have rights to that portion of the site, you have been logged off!")
            return redirect('accounts:logout')
        return super(DisciplineList, self).get(*args, **kwargs)

    def get_queryset(self):
        return  Discipline.objects.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['config_menu'] = '--active'
        context['config_menu_open'] = 'side-menu__sub-open'
        context['discipline_menu_open'] = '--active'
        context['wards'] = Ward.objects.all()
        return context

@login_required()
def discipline_add(request):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        form = DisciplineForm(request.POST)
        if form.is_valid():
            form.save()                       

            messages.success(request,"Successfully added discipline details")
        else:
            messages.warning(request,form.errors)

        return redirect('configurable:discipline_list')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')

@login_required()
def discipline_edit(request,pk):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        block_instance = Discipline.objects.get(id = pk)
        form = DisciplineForm(request.POST,instance = block_instance)
        if form.is_valid():
            form.save()
            
            messages.success(request,"Successfully edited discipline")
        else:
            messages.warning(request,form.errors)

        return redirect('configurable:discipline_list')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')

@login_required()
def discipline_delete(request,pk):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        try:
            i_instance = Discipline.objects.get(id = pk)
            i_instance.delete()
            messages.success(request,"Successfully deleted discipline")
        except Exception as e:
            messages.warning(request,f"An error has occurred, please try again - Error: {str(e)}")

        return redirect('configurable:discipline_list')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')
    
    
@login_required()
def discipline_ward_add(request,pk):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        discipline = Discipline.objects.get(id = pk)

        wards = request.POST.getlist('wards[]')
     
        for ward_pk in wards:
          
            ward = Ward.objects.get(id = ward_pk)
            discipline.wards.add(ward)
            discipline.save()
            
        messages.success(request,'Successfully added wards to discipline')

        return redirect('configurable:discipline_list')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')
    
    

@login_required()
def discipline_ward_remove(request,pk,ward_pk):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        discipline = Discipline.objects.get(id = pk)
        ward = Ward.objects.get(id = ward_pk)
        
        discipline.wards.remove(ward)
        discipline.save()
            
        messages.success(request,'Successfully removed ward from discipline')

        return redirect('configurable:discipline_list')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')
    

class WardList(LoginRequiredMixin,ListView):
    template_name = 'configurable/wards.html'
    context_object_name = 'items'

    def get(self, *args, **kwargs):
        if self.request.user.logged_in_role_id != 2 and self.request.user.logged_in_role_id != 1:
            messages.warning(self.request,"You do not have rights to that portion of the site, you have been logged off!")
            return redirect('accounts:logout')
        return super(WardList, self).get(*args, **kwargs)

    def get_queryset(self):
        return  Ward.objects.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['config_menu'] = '--active'
        context['config_menu_open'] = 'side-menu__sub-open'
        context['ward_menu_open'] = '--active'
        return context
    

@login_required()
def ward_add(request):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        form = WardForm(request.POST)
        if form.is_valid():
            ward = form.save(commit = False)
            ward.save()
            
            messages.success(request,"Successfully added ward")
        else:
            messages.warning(request,form.errors)

        return redirect('configurable:ward_list')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')
    

@login_required()
def ward_edit(request,pk):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        ward = Ward.objects.get(id = pk)
        form = WardForm(request.POST,instance=ward)
        if form.is_valid():
            ward = form.save(commit = False)
            ward.save()
            
            messages.success(request,"Successfully edited ward")
        else:
            messages.warning(request,form.errors)

        return redirect('configurable:ward_list')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')
    

@login_required()
def ward_delete(request,pk):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        try:
            i_instance = Ward.objects.get(id = pk)
            i_instance.delete()
            messages.success(request,"Successfully removed ward from discipline")
        except Exception as e:
            messages.warning(request,f"An error has occurred, please try again - Error: {str(e)}")

        return redirect('configurable:ward_list')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')
    
    
class ClinicalProcedureThemeList(LoginRequiredMixin,ListView):
    template_name = 'configurable/clinical_procedures.html'
    context_object_name = 'items'

    def get(self, *args, **kwargs):
        if self.request.user.logged_in_role_id != 2 and self.request.user.logged_in_role_id != 1:
            messages.warning(self.request,"You do not have rights to that portion of the site, you have been logged off!")
            return redirect('accounts:logout')
        return super(ClinicalProcedureThemeList, self).get(*args, **kwargs)

    def get_queryset(self):
        return  ClinicalProcedureTheme.objects.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['config_menu'] = '--active'
        context['config_menu_open'] = 'side-menu__sub-open'
        context['clinical_procedure_menu_open'] = '--active'
        return context

@login_required()
def clinical_procedure_add(request):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        form = ClinicalProcedureThemeForm(request.POST)
        if form.is_valid():
            form.save()                       

            messages.success(request,"Successfully added Clinical Procedure Theme details")
        else:
            messages.warning(request,form.errors)

        return redirect('configurable:clinical_procedure_list')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')

@login_required()
def clinical_procedure_edit(request,pk):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        block_instance = ClinicalProcedureTheme.objects.get(id = pk)
        form = ClinicalProcedureThemeForm(request.POST,instance = block_instance)
        if form.is_valid():
            form.save()
            
            messages.success(request,"Successfully edited Clinical Procedure Theme")
        else:
            messages.warning(request,form.errors)

        return redirect('configurable:clinical_procedure_list')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')

@login_required()
def clinical_procedure_delete(request,pk):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        try:
            i_instance = ClinicalProcedureTheme.objects.get(id = pk)
            i_instance.delete()
            messages.success(request,"Successfully deleted Clinical Procedure Theme")
        except Exception as e:
            messages.warning(request,f"An error has occurred, please try again - Error: {str(e)}")

        return redirect('configurable:clinical_procedure_list')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')
    

@login_required()
def clinical_procedure_task_add(request,pk):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        theme = ClinicalProcedureTheme.objects.get(id = pk)
        form = ClinicalProcedureThemeTaskForm(request.POST)
        if form.is_valid():
            task = form.save(commit = False)
            task.theme = theme
            task.save()
            
            messages.success(request,"Successfully added task")
        else:
            messages.warning(request,form.errors)

        return redirect('configurable:clinical_procedure_list')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')
    

@login_required()
def clinical_procedure_task_edit(request,pk):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        task = ClinicalProcedureThemeTask.objects.get(id = pk)
        form = ClinicalProcedureThemeTaskForm(request.POST,instance=task)
        if form.is_valid():
            task = form.save(commit = False)
            task.save()
            
            messages.success(request,"Successfully edited task")
        else:
            messages.warning(request,form.errors)

        return redirect('configurable:clinical_procedure_list')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')
    

@login_required()
def clinical_procedure_task_delete(request,pk):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        try:
            i_instance = Ward.objects.get(id = pk)
            i_instance.delete()
            messages.success(request,"Successfully removed task from theme")
        except Exception as e:
            messages.warning(request,f"An error has occurred, please try again - Error: {str(e)}")

        return redirect('configurable:clinical_procedure_list')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')
    


class ClinicalProcedureThemeTaskAssessmentList(LoginRequiredMixin,ListView):
    template_name = 'configurable/clinical_procedures_assessments.html'
    context_object_name = 'items'

    def get(self, *args, **kwargs):
        if self.request.user.logged_in_role_id != 2 and self.request.user.logged_in_role_id != 1:
            messages.warning(self.request,"You do not have rights to that portion of the site, you have been logged off!")
            return redirect('accounts:logout')
        return super(ClinicalProcedureThemeTaskAssessmentList, self).get(*args, **kwargs)

    def get_queryset(self):
        return  ClinicalProcedureThemeTaskAssessment.objects.filter(task = self.kwargs['pk'])

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['config_menu'] = '--active'
        context['config_menu_open'] = 'side-menu__sub-open'
        context['clinical_procedure_menu_open'] = '--active'
        context['task'] = ClinicalProcedureThemeTask.objects.get(id = self.kwargs['pk'])
        #get all tasks that have assessments
        context['tasks'] = tasks_with_assessments = ClinicalProcedureThemeTask.objects.filter(assessments__isnull=False).distinct()
        return context

@login_required()
def clinical_procedure_task_assessment_add(request,pk):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        form = ClinicalProcedureThemeTaskAssessmentForm(request.POST)
        if form.is_valid():
            assessment = form.save(commit=False)  
            assessment.task_id = pk  
            assessment.save()                   

            messages.success(request,"Successfully added Assessment details")
        else:
            messages.warning(request,form.errors)

        return redirect('configurable:clinical_procedure_assessment_list',pk=pk)
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')



@login_required()
def clinical_procedure_task_assessment_copy(request,pk):
    
    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:
        
        source_task = ClinicalProcedureThemeTask.objects.get(id = request.POST['task'])

        for assessment in source_task.assessments.all():
            ClinicalProcedureThemeTaskAssessment.objects.create(
                task_id=pk,
                question=assessment.question,
                question_type=assessment.question_type,
                number=assessment.number,
                penalty=assessment.penalty
            )
            
        messages.success(request,'Successfully copied assessments.')

        return redirect('configurable:clinical_procedure_assessment_list',pk=pk)
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')
    
@login_required()
def clinical_procedure_assessment_edit(request,pk,assessment_pk):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        block_instance = ClinicalProcedureThemeTaskAssessment.objects.get(id = assessment_pk)
        form = ClinicalProcedureThemeTaskAssessmentForm(request.POST,instance = block_instance)
        if form.is_valid():
            form.save()
            
            messages.success(request,"Successfully edited Assessment")
        else:
            messages.warning(request,form.errors)

        return redirect('configurable:clinical_procedure_assessment_list',pk=pk)
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')

@login_required()
def clinical_procedure_assessment_delete(request,pk,assessment_pk):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        try:
            i_instance = ClinicalProcedureThemeTaskAssessment.objects.get(id = assessment_pk)
            i_instance.delete()
            messages.success(request,"Successfully deleted Assessment")
        except Exception as e:
            messages.warning(request,f"An error has occurred, please try again - Error: {str(e)}")

        return redirect('configurable:clinical_procedure_assessment_list',pk=pk)
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')


@login_required()
def clinical_procedure_assessment_bulk_upload(request,pk):

    '''
    Bulk upload of clinical procedures
    
    COL 0: Question Number
    COL 1: Critical Mark
    COL 2: Question Text
    COL 3: Question Type
    COL 4: Penalty Amount
    '''

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        file_in_memory = request.FILES['assessment_csv']
        decoded_file = file_in_memory.read().decode('utf-8-sig')
        csv_data = csv.reader(StringIO(decoded_file), delimiter=',')
        x=0

        task = ClinicalProcedureThemeTask.objects.get(id = pk)

        for cols in csv_data:
                   
            try:     
                print(cols)          
                   
                if cols[2]: 
                    
                    penalty = 0
                    
                    if cols[1] == "*":
                        penalty = 10
                    elif cols[1] == "#":
                        penalty = 50
                    
                    ClinicalProcedureThemeTaskAssessment.objects.create(
                        task=task,
                        question=cols[2],
                        question_type=cols[3],
                        number=cols[0],
                        penalty=penalty
                    ) 


            except Exception as e:
                messages.warning(request,f"COL: {x} - {str(e)}")

            x = x + 1

            return redirect('configurable:clinical_procedure_assessment_list',pk=pk)
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')
    

@login_required()
def clinical_procedure_assessment_bulk_upload_excel(request, pk):
    '''
    Bulk upload of clinical procedures from Excel

    COL 0: Question Number
    COL 1: Critical Mark
    COL 2: Question Text
    COL 3: Question Type
    COL 4: Penalty Amount (optional)
    '''

    if request.user.logged_in_role_id in [1, 2]:
        try:
            file_in_memory = request.FILES['assessment_csv']  # Rename this to 'assessment_excel' if you want
            wb = load_workbook(filename=BytesIO(file_in_memory.read()), data_only=True)
            sheet = wb.active
            
            
            task = ClinicalProcedureThemeTask.objects.get(id=pk)
            
            ClinicalProcedureThemeTaskAssessment.objects.filter(task = task).delete()
            
            x = 0

            for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
                try:
                    question_number = row[0]
                    critical_mark = row[1]
                    question_text = row[2]
                    question_type = row[3]

                    if question_text:
                        penalty = 0
                        if critical_mark == "*":
                            penalty = 10
                        elif critical_mark == "#":
                            penalty = 50
                        

                        ClinicalProcedureThemeTaskAssessment.objects.create(
                            task=task,
                            question=question_text,
                            question_type=question_type,
                            number=question_number,
                            penalty=penalty
                        )
                except Exception as inner_e:
                    messages.warning(request, f"Row {x + 2}: {str(inner_e)}")

                x += 1

            messages.success(request, "Bulk upload completed successfully.")
        except Exception as e:
            messages.error(request, f"Upload failed: {str(e)}")

        return redirect('configurable:clinical_procedure_assessment_list', pk=pk)

    else:
        messages.warning(request, "You do not have rights to that portion of the site. You have been logged off!")
        return redirect('accounts:logout')

    
class ShiftList(LoginRequiredMixin,ListView):
    template_name = 'configurable/shifts.html'
    context_object_name = 'items'

    def get(self, *args, **kwargs):
        if self.request.user.logged_in_role_id != 2 and self.request.user.logged_in_role_id != 1:
            messages.warning(self.request,"You do not have rights to that portion of the site, you have been logged off!")
            return redirect('accounts:logout')
        return super(ShiftList, self).get(*args, **kwargs)

    def get_queryset(self):
        return  ShiftType.objects.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['config_menu'] = '--active'
        context['config_menu_open'] = 'side-menu__sub-open'
        context['shift_menu_open'] = '--active'
        return context

@login_required()
def shift_add(request):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        form = ShiftForm(request.POST)
        if form.is_valid():
            form.save()                       

            messages.success(request,"Successfully added Shift")
        else:
            messages.warning(request,form.errors)

        return redirect('configurable:shift_list')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')

@login_required()
def shift_edit(request,pk):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        shift_instance = ShiftType.objects.get(id = pk)
        form = ShiftForm(request.POST,instance = shift_instance)
        if form.is_valid():
            form.save()
            
            messages.success(request,"Successfully edited Progarmme Block")
        else:
            messages.warning(request,form.errors)

        return redirect('configurable:shift_list')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')

@login_required()
def shift_delete(request,pk):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        try:
            i_instance = ShiftType.objects.get(id = pk)
            i_instance.delete()
            messages.success(request,"Successfully deleted Progarmme Block")
        except Exception as e:
            messages.warning(request,f"An error has occurred, please try again - Error: {str(e)}")

        return redirect('configurable:shift_list')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')
    

class WILScaleQuestionList(LoginRequiredMixin,ListView):
    template_name = 'configurable/wil_questions.html'
    context_object_name = 'items'

    def get(self, *args, **kwargs):
        if self.request.user.logged_in_role_id != 2 and self.request.user.logged_in_role_id != 1:
            messages.warning(self.request,"You do not have rights to that portion of the site, you have been logged off!")
            return redirect('accounts:logout')
        return super(WILScaleQuestionList, self).get(*args, **kwargs)

    def get_queryset(self):
        return  WILScaleQuestion.objects.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['config_menu'] = '--active'
        context['config_menu_open'] = 'side-menu__sub-open'
        context['wil_scale_menu_open'] = '--active'
        return context
    

@login_required()
def wil_scale_add(request):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        form = WILScaleQuestionForm(request.POST)
        if form.is_valid():
            wil = form.save(commit = False)
            wil.save()
            
            messages.success(request,"Successfully added Question")
        else:
            messages.warning(request,form.errors)

        return redirect('configurable:wil_scale_list')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')
    

@login_required()
def wil_scale_edit(request,pk):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        wil = WILScaleQuestion.objects.get(id = pk)
        form = WILScaleQuestionForm(request.POST,instance=wil)
        if form.is_valid():
            wil = form.save(commit = False)
            wil.save()
            
            messages.success(request,"Successfully edited")
        else:
            messages.warning(request,form.errors)

        return redirect('configurable:wil_scale_list')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')
    

@login_required()
def wil_scale_delete(request,pk):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        try:
            i_instance = WILScaleQuestion.objects.get(id = pk)
            i_instance.delete()
            messages.success(request,"Successfully removed")
        except Exception as e:
            messages.warning(request,f"An error has occurred, please try again - Error: {str(e)}")

        return redirect('configurable:wil_scale_list')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')


def ajax_fetch_discipline_wards(request):
    discipline_id = request.GET.get('discipline_id', None)
 
    try:
        discipline_check = Discipline.objects.filter(id = discipline_id)
    
        if discipline_check.exists():
            discipline = discipline_check.first()
          
            
            info = []
            if discipline.wards.count() > 0:
                for x in discipline.wards.all():
                    info.append({'id':x.id,'ward':x.ward})
                    
                data = {
                    'valid':1,
                    'discipline_id':discipline.id,
                }
                data['info'] = info
            else:
                data = {
                    'valid':0,
                    'message':'No wards'
                }    
        else:
            data = {
                'valid':0,
                'message':'No wards'
            }
    except Exception as e:
        data = {
            'valid':2,
            'message':'str(e)'
        }

    return JsonResponse(data)


class RegistrationBlockList(LoginRequiredMixin,ListView):
    template_name = 'configurable/registration_blocks.html'
    context_object_name = 'items'

    def get(self, *args, **kwargs):
        if self.request.user.logged_in_role_id != 2 and self.request.user.logged_in_role_id != 1:
            messages.warning(self.request,"You do not have rights to that portion of the site, you have been logged off!")
            return redirect('accounts:logout')
        return super(RegistrationBlockList, self).get(*args, **kwargs)

    def get_queryset(self):
        return  RegistrationBlockCode.objects.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['config_menu'] = '--active'
        context['config_menu_open'] = 'side-menu__sub-open'
        context['registration_block_menu_open'] = '--active'
        return context

@login_required()
def registration_block_add(request):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        form = RegistrationBlockCodeForm(request.POST)
        if form.is_valid():
            form.save()                       

            messages.success(request,"Successfully added Registration Block details")
        else:
            messages.warning(request,form.errors)

        return redirect('configurable:registration_block_list')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')

@login_required()
def registration_block_edit(request,pk):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        block_instance = RegistrationBlockCode.objects.get(id = pk)
        form = RegistrationBlockCodeForm(request.POST,instance = block_instance)
        if form.is_valid():
            form.save()
            
            messages.success(request,"Successfully edited Registration Block")
        else:
            messages.warning(request,form.errors)

        return redirect('configurable:registration_block_list')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')

@login_required()
def registration_block_delete(request,pk):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        try:
            i_instance = RegistrationBlockCode.objects.get(id = pk)
            i_instance.delete()
            messages.success(request,"Successfully deleted Registration Block")
        except Exception as e:
            messages.warning(request,f"An error has occurred, please try again - Error: {str(e)}")

        return redirect('configurable:registration_block_list')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')
    


class RegisterCategoryList(LoginRequiredMixin,ListView):
    template_name = 'configurable/register_categories.html'
    context_object_name = 'items'

    def get(self, *args, **kwargs):
        if self.request.user.logged_in_role_id != 2 and self.request.user.logged_in_role_id != 1:
            messages.warning(self.request,"You do not have rights to that portion of the site, you have been logged off!")
            return redirect('accounts:logout')
        return super(RegisterCategoryList, self).get(*args, **kwargs)

    def get_queryset(self):
        return  RegisterCategory.objects.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['config_menu'] = '--active'
        context['config_menu_open'] = 'side-menu__sub-open'
        context['register_category_menu_open'] = '--active'
        return context

@login_required()
def register_category_add(request):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        form = RegisterCategoryForm(request.POST)
        if form.is_valid():
            form.save()                       

            messages.success(request,"Successfully added category")
        else:
            messages.warning(request,form.errors)

        return redirect('configurable:register_category_list')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')

@login_required()
def register_category_edit(request,pk):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        block_instance = RegisterCategory.objects.get(id = pk)
        form = RegisterCategoryForm(request.POST,instance = block_instance)
        if form.is_valid():
            form.save()
            
            messages.success(request,"Successfully edited Register Category")
        else:
            messages.warning(request,form.errors)

        return redirect('configurable:register_category_list')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')

@login_required()
def register_category_delete(request,pk):

    if request.user.logged_in_role_id == 2 or request.user.logged_in_role_id == 1:

        try:
            i_instance = RegisterCategory.objects.get(id = pk)
            i_instance.delete()
            messages.success(request,"Successfully deleted Register Category")
        except Exception as e:
            messages.warning(request,f"An error has occurred, please try again - Error: {str(e)}")

        return redirect('configurable:register_category_list')
    else:
        messages.warning(request,"You do not have rights to that portion of the site, you have been logged off!")
        return redirect('accounts:logout')
    

class DoseList(LoginRequiredMixin,ListView):
    template_name = 'configurable/dose.html'
    context_object_name = 'doses'
    model = VaccinationDose

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['config_menu'] = '--active'
        context['config_menu_open'] = 'side-menu__sub-open'
        context['dose_menu_open'] = '--active'
        return context


@login_required()
def add_dose(request):

    form = VaccinationDoseForm(request.POST)
    if form.is_valid():
        dose = form.save()
        messages.success(request,'Successfully added Vaccination Dose')
    else:
        messages.warning(request,form.errors)

    return redirect('configurable:doses')

@login_required()
def edit_dose(request,pk):

    dose_instance = VaccinationDose.objects.get(id=pk)
    form = VaccinationDoseForm(request.POST,instance=dose_instance)

    if form.is_valid():
        dose = form.save(commit=True)
        messages.success(request,'Successfully edited Vaccination Dose')
    else:
        messages.warning(request,form.errors)

    return redirect('configurable:doses')

@login_required()
def delete_dose(request,pk):

    try:
        dose_instance = VaccinationDose.objects.get(id=pk)
        dose_instance.delete()
        messages.success(request, 'Successfully deleted Vaccination Dose')
    except Exception as e:
        messages.warning(request, str(e))

    return redirect('configurable:doses')


